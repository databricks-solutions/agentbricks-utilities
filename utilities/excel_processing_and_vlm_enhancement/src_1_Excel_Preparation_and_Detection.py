# Databricks notebook source
# MAGIC %md
# MAGIC # Excel Preparation and Visual Element Detection
# MAGIC
# MAGIC First stage of Excel processing pipeline:
# MAGIC - Scans Excel files from source volume
# MAGIC - Detects visual elements (charts, images, shapes)
# MAGIC - Adjusts page setup for PDF conversion
# MAGIC - Saves PDF-ready Excel files to prepared subfolder
# MAGIC - Creates metadata table for tracking
# MAGIC
# MAGIC **Output**: Metadata table and PDF-ready Excel files
# MAGIC **Next Step**: Run notebook 2 for PDF conversion

# COMMAND ----------

# MAGIC %md
# MAGIC ## 1. Environment Setup

# COMMAND ----------

# MAGIC %pip install openpyxl --quiet

# COMMAND ----------

# MAGIC %md
# MAGIC ## 2. Parameter Configuration

# COMMAND ----------

# Create widgets for job parameters
dbutils.widgets.text("catalog", "main", "1. Catalog Name")
dbutils.widgets.text("schema", "default", "2. Schema Name")
dbutils.widgets.text("source_volume", "source_excel", "3. Source Volume Name")
dbutils.widgets.text("prepared_volume", "excel_prepared", "4. Prepared Excel Volume Name")
dbutils.widgets.text("metadata_table", "excel_metadata", "5. Metadata Table Name")
dbutils.widgets.dropdown("paper_size", "A4", ["A4", "A3"], "6. Paper Size")
dbutils.widgets.text("source_subfolder", "", "7. Source Subfolder (optional)")
dbutils.widgets.dropdown("process_mode", "overwrite", ["overwrite", "append"], "8. Process Mode")

# Get parameter values
catalog = dbutils.widgets.get("catalog")
schema = dbutils.widgets.get("schema")
source_volume = dbutils.widgets.get("source_volume")
prepared_volume = dbutils.widgets.get("prepared_volume")
metadata_table = dbutils.widgets.get("metadata_table")
paper_size = dbutils.widgets.get("paper_size")
source_subfolder = dbutils.widgets.get("source_subfolder")
process_mode = dbutils.widgets.get("process_mode")

# Construct volume paths
source_volume_path = f"/Volumes/{catalog}/{schema}/{source_volume}"
if source_subfolder:
    source_volume_path = f"{source_volume_path}/{source_subfolder.strip('/')}"

prepared_volume_path = f"/Volumes/{catalog}/{schema}/{prepared_volume}/pdf_ready_excel"
metadata_table_name = f"{catalog}.{schema}.{metadata_table}"

# Print configuration
print("=" * 80)
print("EXCEL PREPARATION & DETECTION CONFIGURATION")
print("=" * 80)
print(f"Source Volume Path:     {source_volume_path}")
print(f"Prepared Excel Path:    {prepared_volume_path}")
print(f"Metadata Table:         {metadata_table_name}")
print(f"Paper Size:             {paper_size}")
print(f"Process Mode:           {process_mode}")
print("=" * 80)

# COMMAND ----------

import json
json.dumps(dbutils.widgets.getAll())

# COMMAND ----------

# MAGIC %md
# MAGIC ## 3. Import Libraries

# COMMAND ----------

import os
import json
import tempfile
from pathlib import Path
from datetime import datetime
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Tuple, Optional

from pyspark.sql.functions import col, current_timestamp
from pyspark.sql.types import (
    StructType, StructField, StringType, 
    BooleanType, IntegerType, LongType, TimestampType
)
from pyspark.sql import Row

import openpyxl
from openpyxl import load_workbook

print("All libraries imported successfully")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 4. Visual Element Detection Functions

# COMMAND ----------

def detect_visual_elements(file_path: str) -> Dict:
    """
    Detect visual elements in Excel file across all sheets.
    
    Args:
        file_path: Full path to the Excel file
    Returns:
        Dictionary with detection results and metadata
    """
    result = {
        'file_path': file_path,
        'file_name': os.path.basename(file_path),
        'has_chart': False,
        'has_image': False,
        'has_shape': False,
        'total_sheets': 0,
        'total_charts': 0,
        'total_images': 0,
        'total_tables': 0,
        'sheet_details': [],
        'detection_error': None,
        'detection_status': 'success'
    }
    
    try:
        # Load workbook
        wb = load_workbook(file_path, data_only=False)
        result['total_sheets'] = len(wb.sheetnames)
        
        # Iterate through all sheets
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            sheet_info = {
                'sheet_name': sheet_name,
                'sheet_index': wb.sheetnames.index(sheet_name),
                'has_chart': False,
                'has_image': False,
                'chart_count': 0,
                'image_count': 0,
                'table_count': 0,
                'chart_types': []
            }
            
            # Detect charts
            if hasattr(ws, '_charts') and len(ws._charts) > 0:
                sheet_info['has_chart'] = True
                sheet_info['chart_count'] = len(ws._charts)
                result['has_chart'] = True
                result['total_charts'] += len(ws._charts)
                
                # Get chart types
                for chart in ws._charts:
                    chart_type = type(chart).__name__
                    sheet_info['chart_types'].append(chart_type)
            
            # Detect images
            if hasattr(ws, '_images') and len(ws._images) > 0:
                sheet_info['has_image'] = True
                sheet_info['image_count'] = len(ws._images)
                result['has_image'] = True
                result['total_images'] += len(ws._images)
            
            result['sheet_details'].append(sheet_info)
        
        # Detect shapes from XML (additional check)
        result['has_shape'] = detect_shapes_from_xml(file_path)
        
        wb.close()
        
    except Exception as e:
        result['detection_error'] = str(e)
        result['detection_status'] = 'failed'
    
    return result


def detect_shapes_from_xml(file_path: str) -> bool:
    """
    Parse Excel XML to detect shapes and drawing objects.
    
    Args:
        file_path: Full path to Excel file
    Returns:
        Boolean indicating presence of shapes
    """
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            # Check for drawing files in Excel XML structure
            drawing_files = [
                f for f in zip_ref.namelist() 
                if f.startswith('xl/drawings/') and f.endswith('.xml')
            ]
            
            if drawing_files:
                for drawing_file in drawing_files:
                    xml_content = zip_ref.read(drawing_file)
                    root = ET.fromstring(xml_content)
                    
                    # Check for shape elements in XML
                    namespaces = {
                        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
                    }
                    shapes = root.findall('.//xdr:sp', namespaces)
                    
                    if shapes:
                        return True
            
            return False
            
    except Exception:
        return False


def get_file_info(file_path: str) -> Dict:
    """
    Get file size and modification time.
    
    Args:
        file_path: Full path to file
    Returns:
        Dictionary with file metadata
    """
    try:
        file_stat = os.stat(file_path)
        return {
            'file_size_bytes': file_stat.st_size,
            'file_size_mb': round(file_stat.st_size / (1024 * 1024), 2),
            'last_modified': datetime.fromtimestamp(file_stat.st_mtime).isoformat()
        }
    except Exception as e:
        return {
            'file_size_bytes': None,
            'file_size_mb': None,
            'last_modified': None,
            'file_info_error': str(e)
        }

# COMMAND ----------

# MAGIC %md
# MAGIC ## 5. Excel Page Setup Functions

# COMMAND ----------

def adjust_page_setup_for_pdf(
    source_path: str,
    output_path: str,
    paper_size: str = "A4"
) -> Tuple[str, bool, Optional[str]]:
    """
    Adjust Excel page setup for PDF conversion.
    
    Adjustments: landscape orientation, fit to 1 page wide, reduced margins, paper size.
    Uses temp storage to avoid volume file locking issues.
    
    Args:
        source_path: Path to source Excel file
        output_path: Path for PDF-ready Excel file
        paper_size: A3 or A4
    Returns:
        Tuple of (output_path, success, error_message)
    """
    temp_file = None
    
    try:
        # Load workbook from source
        wb = load_workbook(source_path)

        # Track worksheet processing errors
        sheet_errors = []
        
        # Apply page setup to all worksheets
        for ws in wb.worksheets:
            try:
                # Skip if worksheet is None or doesn't have required attributes
                if ws is None:
                    sheet_errors.append(f"Worksheet is None")
                    continue
                
                # Clear print area
                if hasattr(ws, 'print_area'):
                    ws.print_area = None
                
                # Set landscape orientation
                if hasattr(ws, 'page_setup') and hasattr(ws, 'ORIENTATION_LANDSCAPE'):
                    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
                    
                    # Fit to 1 page wide, unlimited height
                    ws.page_setup.fitToPage = True
                    ws.page_setup.fitToWidth = 1
                    ws.page_setup.fitToHeight = 0  # 0 = unlimited
                    
                    # Set paper size
                    if paper_size.upper() == "A3" and hasattr(ws, 'PAPERSIZE_A3'):
                        ws.page_setup.paperSize = ws.PAPERSIZE_A3
                    elif hasattr(ws, 'PAPERSIZE_A4'):
                        ws.page_setup.paperSize = ws.PAPERSIZE_A4
                
                # Reduce margins for more content space
                if hasattr(ws, 'page_margins'):
                    ws.page_margins.left = 0.3
                    ws.page_margins.right = 0.3
                    ws.page_margins.top = 0.3
                    ws.page_margins.bottom = 0.3
                    ws.page_margins.header = 0.2
                    ws.page_margins.footer = 0.2
                    
            except AttributeError as e:
                # Some worksheets might not support page setup (e.g., chart sheets)
                sheet_name = getattr(ws, 'title', 'Unknown')
                sheet_errors.append(f"Sheet '{sheet_name}': {str(e)}")
                continue
            except Exception as e:
                # Catch any other worksheet-specific errors
                sheet_name = getattr(ws, 'title', 'Unknown')
                sheet_errors.append(f"Sheet '{sheet_name}': {str(e)}")
                continue
        
        # Create temporary file for saving (openpyxl can't save directly to volumes)
        temp_fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)  # Close file descriptor, we'll use the path
        
        # Save to temporary local file
        wb.save(temp_file)
        wb.close()
        
        # Copy from temp to volume using dbutils
        dbutils.fs.cp(f"file:{temp_file}", output_path, recurse=True)
        
        # Clean up temp file
        if os.path.exists(temp_file):
            os.remove(temp_file)
        
        # Return success even if some sheets had issues (file was still saved)
        if sheet_errors:
            warning_msg = f"Completed with warnings: {'; '.join(sheet_errors[:3])}"
            if len(sheet_errors) > 3:
                warning_msg += f" (and {len(sheet_errors)-3} more)"
            return (output_path, True, warning_msg)

        return (output_path, True, None)
        
    except Exception as e:
        # Clean up temp file on error
        if temp_file and os.path.exists(temp_file):
            try:
                os.remove(temp_file)
            except:
                pass
        return (output_path, False, str(e))

# COMMAND ----------

# MAGIC %md
# MAGIC ## 6. File Processing Function

# COMMAND ----------

def process_excel_file(
    source_file_path: str,
    prepared_dir: str,
    paper_size: str
) -> Dict:
    """
    Process single Excel file: detect elements and create PDF-ready version.
    
    Args:
        source_file_path: Path to source Excel file
        prepared_dir: Directory for PDF-ready files
        paper_size: A3 or A4
    Returns:
        Dictionary with processing results
    """
    start_time = datetime.now()
    
    # Initialize result dictionary
    result = {
        'source_file_path': source_file_path,
        'source_file_name': os.path.basename(source_file_path),
        'pdf_ready_excel_path': None,
        'pdf_file_path': None,  # Will be populated by notebook 2
        'processing_start_time': start_time.isoformat(),
        'processing_end_time': None,
        'processing_duration_seconds': None,
        'preparation_status': 'success',
        'pdf_conversion_status': 'pending',  # Will be updated by notebook 2
        'errors': []
    }
    
    try:
        # Step 1: Get file info
        file_info = get_file_info(source_file_path)
        result.update(file_info)
        
        # Step 2: Detect visual elements
        visual_info = detect_visual_elements(source_file_path)
        result.update({
            'has_chart': visual_info['has_chart'],
            'has_image': visual_info['has_image'],
            'has_shape': visual_info['has_shape'],
            'has_visual_element': (
                visual_info['has_chart'] or 
                visual_info['has_image'] or 
                visual_info['has_shape']
            ),
            'total_sheets': visual_info['total_sheets'],
            'total_charts': visual_info['total_charts'],
            'total_images': visual_info['total_images'],
            'total_tables': visual_info['total_tables'],
            'sheet_details_json': json.dumps(visual_info['sheet_details'])
        })
        
        if visual_info['detection_error']:
            result['errors'].append(f"Visual detection: {visual_info['detection_error']}")
        
        # Step 3: Create PDF-ready Excel file
        source_stem = Path(source_file_path).stem
        source_ext = Path(source_file_path).suffix
        pdf_ready_filename = f"{source_stem}_pdfready{source_ext}"
        pdf_ready_path = os.path.join(prepared_dir, pdf_ready_filename)
        
        pdf_ready_result = adjust_page_setup_for_pdf(
            source_file_path,
            pdf_ready_path,
            paper_size
        )
        
        if pdf_ready_result[1]:  # Success
            result['pdf_ready_excel_path'] = pdf_ready_result[0]
        else:
            result['errors'].append(f"Page setup: {pdf_ready_result[2]}")
            result['preparation_status'] = 'failed'
        
    except Exception as e:
        result['errors'].append(f"Unexpected error: {str(e)}")
        result['preparation_status'] = 'failed'
    
    # Calculate processing duration
    end_time = datetime.now()
    result['processing_end_time'] = end_time.isoformat()
    result['processing_duration_seconds'] = (end_time - start_time).total_seconds()
    result['errors_json'] = json.dumps(result['errors'])
    
    return result

# COMMAND ----------

# MAGIC %md
# MAGIC ## 7. Main Batch Processing Execution

# COMMAND ----------

# Create output directory
os.makedirs(prepared_volume_path, exist_ok=True)
print("Output directory created/verified")
print()

# Query Excel files from bronze_documents table
print("Querying Excel files from bronze layer...")
try:
    excel_files_df = spark.sql(f"""
        SELECT document_id, path, content, file_name, file_size_bytes
        FROM {catalog}.{schema}.bronze_documents
        WHERE file_type = 'spreadsheet'
        AND (processing_status = 'ingested' OR processing_status = 'conversion_failed')
        ORDER BY ingestion_timestamp DESC
    """)
    excel_files_data = excel_files_df.collect()
    excel_files = [row.path.replace('dbfs:', '') for row in excel_files_data]
    # Store mapping of path to document_id for later updates
    path_to_doc_id = {row.path.replace('dbfs:', ''): row.document_id for row in excel_files_data}
    print(f"Found {len(excel_files)} Excel file(s)")
except Exception as e:
    print(f"ERROR: Failed to query bronze_documents table")
    print(f"Error: {str(e)}")
    dbutils.notebook.exit(f"Failed to access bronze_documents: {str(e)}")

if len(excel_files) == 0:
    print("WARNING: No Excel files found to process")
    dbutils.notebook.exit("No Excel files found")

print()
print("=" * 80)
print("STARTING BATCH PROCESSING")
print("=" * 80)

# Process each file and collect results
processing_results = []

for idx, file_path in enumerate(excel_files, 1):
    file_name = os.path.basename(file_path)
    print(f"\n[{idx}/{len(excel_files)}] Processing: {file_name}")
    print("-" * 80)
    
    result = process_excel_file(
        source_file_path=file_path,
        prepared_dir=prepared_volume_path,
        paper_size=paper_size
    )
    
    processing_results.append(result)
    
    # Print summary for this file
    status_emoji = "✓" if result['preparation_status'] == 'success' else "✗"
    print(f"Status: {status_emoji} {result['preparation_status'].upper()}")
    print(f"Duration: {result['processing_duration_seconds']:.2f}s")
    
    if result['has_visual_element']:
        print(f"Visual Elements: Charts={result['total_charts']}, "
              f"Images={result['total_images']}, Tables={result['total_tables']}")
    
    if result['errors']:
        print(f"Errors: {len(result['errors'])}")
        for error in result['errors']:
            print(f"  - {error}")

print()
print("=" * 80)
print("BATCH PROCESSING COMPLETED")
print("=" * 80)

# COMMAND ----------

# MAGIC %md
# MAGIC ## 8. Create Metadata Table

# COMMAND ----------

# Define schema for metadata table
metadata_schema = StructType([
    # Source file information
    StructField("source_file_path", StringType(), False),
    StructField("source_file_name", StringType(), False),
    StructField("file_size_bytes", LongType(), True),
    StructField("file_size_mb", StringType(), True),
    StructField("last_modified", StringType(), True),
    
    # Output file paths
    StructField("pdf_ready_excel_path", StringType(), True),
    StructField("pdf_file_path", StringType(), True),
    StructField("pdf_file_size_bytes", LongType(), True),
    StructField("pdf_file_size_mb", StringType(), True),
    
    # Visual elements detection
    StructField("has_chart", BooleanType(), True),
    StructField("has_image", BooleanType(), True),
    StructField("has_shape", BooleanType(), True),
    StructField("has_visual_element", BooleanType(), True),
    
    # Visual elements counts
    StructField("total_sheets", IntegerType(), True),
    StructField("total_charts", IntegerType(), True),
    StructField("total_images", IntegerType(), True),
    StructField("total_tables", IntegerType(), True),
    
    # Sheet-level details (JSON)
    StructField("sheet_details_json", StringType(), True),
    
    # Processing information
    StructField("processing_start_time", StringType(), True),
    StructField("processing_end_time", StringType(), True),
    StructField("processing_duration_seconds", StringType(), True),
    StructField("preparation_status", StringType(), True),
    StructField("pdf_conversion_status", StringType(), True),
    StructField("errors_json", StringType(), True),
    
    # Job metadata
    StructField("paper_size", StringType(), True),
    StructField("preparation_timestamp", TimestampType(), True),
    StructField("pdf_conversion_timestamp", TimestampType(), True)
])

# Convert results to Spark DataFrame
print("Creating metadata DataFrame...")

metadata_rows = []
for result in processing_results:
    row = Row(
        source_file_path=result.get('source_file_path'),
        source_file_name=result.get('source_file_name'),
        file_size_bytes=result.get('file_size_bytes'),
        file_size_mb=str(result.get('file_size_mb')) if result.get('file_size_mb') is not None else None,
        last_modified=result.get('last_modified'),
        pdf_ready_excel_path=result.get('pdf_ready_excel_path'),
        pdf_file_path=result.get('pdf_file_path'),
        pdf_file_size_bytes=result.get('pdf_file_size_bytes'),
        pdf_file_size_mb=str(result.get('pdf_file_size_mb')) if result.get('pdf_file_size_mb') is not None else None,
        has_chart=result.get('has_chart'),
        has_image=result.get('has_image'),
        has_shape=result.get('has_shape'),
        has_visual_element=result.get('has_visual_element'),
        total_sheets=result.get('total_sheets'),
        total_charts=result.get('total_charts'),
        total_images=result.get('total_images'),
        total_tables=result.get('total_tables'),
        sheet_details_json=result.get('sheet_details_json'),
        processing_start_time=result.get('processing_start_time'),
        processing_end_time=result.get('processing_end_time'),
        processing_duration_seconds=str(result.get('processing_duration_seconds')) if result.get('processing_duration_seconds') is not None else None,
        preparation_status=result.get('preparation_status'),
        pdf_conversion_status=result.get('pdf_conversion_status'),
        errors_json=result.get('errors_json'),
        paper_size=paper_size,
        preparation_timestamp=datetime.now(),
        pdf_conversion_timestamp=None
    )
    metadata_rows.append(row)

metadata_df = spark.createDataFrame(metadata_rows, schema=metadata_schema)

# Write to Delta table
print(f"Writing metadata to table: {metadata_table_name}")
metadata_df.write \
    .format("delta") \
    .mode(process_mode) \
    .option("mergeSchema", "true") \
    .saveAsTable(metadata_table_name)

print(f"Metadata table created/updated successfully")

# Update bronze_documents table with preparation status
print("\nUpdating bronze layer with preparation status...")
try:
    for result in processing_results:
        file_path = result['source_file_path']
        if file_path in path_to_doc_id:
            document_id = path_to_doc_id[file_path]
            status = 'prepared' if result['preparation_status'] == 'success' else 'preparation_failed'
            
            # Only update status here; converted_pdf_path will be updated by src_2 after PDF conversion
            spark.sql(f"""
                UPDATE {catalog}.{schema}.bronze_documents
                SET processing_status = '{status}'
                WHERE document_id = '{document_id}'
            """)
    print("Bronze layer updated successfully")
except Exception as e:
    print(f"WARNING: Failed to update bronze layer: {str(e)}")
    print("Metadata table was created successfully, but bronze status update failed")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 9. Processing Summary

# COMMAND ----------

print("\n" + "=" * 80)
print("PROCESSING SUMMARY")
print("=" * 80)

# Calculate statistics
total_files = len(processing_results)
successful_files = sum(1 for r in processing_results if r['preparation_status'] == 'success')
failed_files = sum(1 for r in processing_results if r['preparation_status'] == 'failed')
files_with_visuals = sum(1 for r in processing_results if r.get('has_visual_element', False))
files_with_charts = sum(1 for r in processing_results if r.get('has_chart', False))
files_with_images = sum(1 for r in processing_results if r.get('has_image', False))
total_charts = sum(r.get('total_charts', 0) for r in processing_results)
total_images = sum(r.get('total_images', 0) for r in processing_results)
total_tables = sum(r.get('total_tables', 0) for r in processing_results)
total_duration = sum(r.get('processing_duration_seconds', 0) for r in processing_results)

print(f"\nTotal Files Processed:        {total_files}")
print(f"Successful Preparations:      {successful_files}")
print(f"Failed Preparations:          {failed_files}")
print()
print(f"Files with Visual Elements:   {files_with_visuals}")
print(f"  - Files with Charts:        {files_with_charts}")
print(f"  - Files with Images:        {files_with_images}")
print()
print(f"Total Charts Detected:        {total_charts}")
print(f"Total Images Detected:        {total_images}")
print()
print(f"Total Processing Time:        {total_duration:.2f} seconds")
print(f"Average Time per File:        {total_duration/total_files:.2f} seconds")
print()
print(f"Output Locations:")
print(f"  PDF-Ready Excel Files:      {prepared_volume_path}")
print(f"  Metadata Table:             {metadata_table_name}")
print()

# Show failed files if any
if failed_files > 0:
    print("\nFailed Files:")
    print("-" * 80)
    for result in processing_results:
        if result['preparation_status'] == 'failed':
            print(f"  • {result['source_file_name']}")
            for error in result['errors']:
                print(f"    - {error}")

print("=" * 80)

# Display metadata table preview
print("\nMetadata Table Preview:")
display(metadata_df.select(
    'source_file_name',
    'has_chart',
    'has_image',
    'has_visual_element',
    'total_charts',
    'total_images',
    'preparation_status',
    'pdf_conversion_status'
))

# COMMAND ----------

# MAGIC %md
# MAGIC ## 10. Next Steps

# COMMAND ----------

print("\n" + "=" * 80)
print("NEXT STEPS")
print("=" * 80)
print()
print("✅ Excel files prepared and metadata collected")
print()
print("📊 Query the metadata table:")
print(f"   SELECT * FROM {metadata_table_name}")
print()
print("🔍 Example filters for PDF conversion:")
print(f"   -- Files with charts")
print(f"   SELECT * FROM {metadata_table_name} WHERE has_chart = true")
print()
print(f"   -- Files with any visual elements")
print(f"   SELECT * FROM {metadata_table_name} WHERE has_visual_element = true")
print()
print(f"   -- Files with images but no charts")
print(f"   SELECT * FROM {metadata_table_name} WHERE has_image = true AND has_chart = false")
print()
print("🚀 Run Notebook 2 (Excel_to_PDF_Conversion) to convert selected files to PDF")
print("=" * 80)

# COMMAND ----------

# MAGIC %md
# MAGIC ## 11. Job Completion

# COMMAND ----------

# Create job summary for return value
job_summary = {
    "status": "completed",
    "total_files": total_files,
    "successful": successful_files,
    "failed": failed_files,
    "files_with_visuals": files_with_visuals,
    "metadata_table": metadata_table_name,
    "prepared_excel_path": prepared_volume_path,
    "processing_time_seconds": round(total_duration, 2)
}

# Return job summary as notebook output
dbutils.notebook.exit(json.dumps(job_summary))
